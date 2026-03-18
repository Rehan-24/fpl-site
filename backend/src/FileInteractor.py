# Imports
from itertools import chain
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import MergedCell
from typing import List, Union

import pandas as pd
import shutil
import openpyxl
import os, time, json

# Class
class FileInteractor():
    def __init__(self, file_prefix: str, data: dict[int, dict]):
        self._file_prefix = file_prefix
        self._data = data

        # Resolve paths relative to backend/
        base_dir = os.path.dirname(__file__)                 # backend/src
        root_dir = os.path.abspath(os.path.join(base_dir, ".."))  # backend
        self._templates_dir = os.path.join(root_dir, "templates")
        self._results_dir   = os.path.join(root_dir, "results")
        self._latest_dir    = os.path.join(self._results_dir, "latest")
        os.makedirs(self._results_dir, exist_ok=True)
        os.makedirs(self._latest_dir, exist_ok=True)

    
    @property
    def data(self) -> dict[int, dict]:
        return self._data

    @property
    def file_prefix(self) -> str:
        return self._file_prefix

    def color_chip_background(self, val: Union[int, str]):
        """
        @param val: The content of the cell
        """
        if 'Available' in str(val):
            color = 'palegreen'
        elif 'GW' in str(val):
            color = 'lightpink'
        else:
            color = 'white'
        return f'background-color: {color}'

    def color_team_value_background(self, val: str):
        """
        @param val: The content of the cell
        """
        val = float(val)
        if val > 102:
            color = 'darkgreen'
        elif val > 100:
            color = 'black'
        elif val == 100:
            color = 'darkgray'
        elif val < 98:
            color = 'red'
        else:
            color = 'black'
        return f'color: {color}'

    def color_delta_background(self, val: str):
        """
        @param val: The content of the cell
        """
        val = float(val)
        if val > 0.1:
            color = 'darkgreen'
        elif val == 0.1:
            color = 'lightgreen'
        elif val == -0.1:
            color = 'lightcoral'
        elif val < -0.1:
            color = 'red'
        else:
            color = 'darkgray'
        return f'color: {color}'

    def color_fdr_background(self, val):
        """
        """
        if '(5)' in val:
            color = 'lightcoral'
        elif '(4)' in val:
            color = 'lightpink'
        elif '(3)' in val:
            color = 'lightgray'
        elif '(2)' in val:
            color = 'palegreen'
        elif '(1)' in val:
            color = 'mediumseagreen'
        else:
            color = 'white'
        return f'background-color: {color}'

    def create_spreadsheet(self, gameweek: int, overwrite_results: bool=False):
        """
        @param overwrite_results: boolean to determine whether or not the contents in the results folder should be overwritten
        """
        template_filepath = os.path.join(self._templates_dir, f'{self.file_prefix}_template_v3.xlsx')
        results_filepath  = os.path.join(self._results_dir,   f'{self.file_prefix}_results_v3.xlsx')
        template_source_filepath = template_filepath if overwrite_results else results_filepath

        primary_column_names = ['Team', 'Owner','Points',
                                'Wins', 'Draws', 'Losses',
                                'GP', 'Games Left',
                                'Score', 'Score Against', 'Plus/Minus',
                                'GW Points on Bench', 'Season Points on Bench',
                                'GW Transfers', 'GW Transfer Hit', 'Total Transfers Made', 'Total Transfer Hit',
                                'Highest Point Total Possible',
                                'Current Team Value',]

        chip_labels_list = ['Triple Captain 1', 'Bench Boost 1', 'Free Hit 1', 'Wildcard 1', 'Triple Captain 2', 'Bench Boost 2', 'Free Hit 2', 'Wildcard 2']
        primary_column_names.extend(chip_labels_list)

        # Generate the dataframe for the primary table
        primary_df = pd.DataFrame(columns = primary_column_names)
        for datum_key in self.data.keys():
            datum = self.data[datum_key]
            data_list = self.create_primary_data_list(datum, gameweek)
            primary_df.loc[len(primary_df), :] = data_list
            
        # create a dataframe sorted by score
        score_df = primary_df.copy()
        score_df.sort_values(by=['Score'], ascending=[False], inplace=True)
        
        # find the top 3 and bottom 3 scores
        top3_scores = score_df.head(3)
        top3_scores = top3_scores.reset_index(drop=True)
        bottom3_scores = score_df.tail(3)
        bottom3_scores = bottom3_scores.reset_index(drop=True)
        
        
        # Generate the gameweek strings
        #gameweek_strs = []
        #for gw_offset in range(1):
        #    next_gw = gameweek + gw_offset
        #    if next_gw > 38:
        #        gameweek_strs.append('-')
        #    else:
        #        gameweek_strs.append(str(next_gw))

        #fdr_column_names = ['Team and Current Difficulty Ranking', 'Owner and Previous DR',
        #                    f'Previous GW{gameweek_strs[0]} Opponent', 'Current Rating',
        #                    f'Next Opponent (GW{gameweek_strs[1]})', 'Rating T-1 Week',
        #                    f'Next Opponent (GW{gameweek_strs[2]})', 'Rating T-2 Weeks',
        #                    f'Next Opponent (GW{gameweek_strs[3]})', 'Rating T-3 Weeks']
        #fdr_column_names = ['Team and Current Difficulty Ranking', 'Owner and Previous DR',
        #                    f'Previous GW{gameweek_strs[0]} Opponent', 'Current Rating',
        #                    f'Next Opponent (GW{gameweek_strs[1]})', 'Rating T-1 Week']

        # Generate the dataframe for the fdr table
        #fdr_df = pd.DataFrame(columns = fdr_column_names)
        #for datum_key in self.data.keys():
        ##    datum = self.data[datum_key]
        #     data_list = self.create_fdr_data_list(datum)
        #    fdr_df.loc[len(fdr_df), :] = data_list

        # Ensure results file exists on first run OR overwrite explicitly
        if overwrite_results or not os.path.exists(results_filepath):
            shutil.copyfile(template_filepath, results_filepath)

        # Copy the Template to a new sheet for the gameweek
        wb = load_workbook(template_source_filepath)
        gameweek_str = f'GW{gameweek}'
        dup_ws = wb.copy_worksheet(wb['Template'])
        dup_ws.title = gameweek_str
        wb.save(results_filepath)

        # Write the primary sheet
        with pd.ExcelWriter(results_filepath, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # Write the primary table
            primary_df.sort_values(by=['Points', 'Score'], ascending=[False, False], inplace=True)
            
            #print check 
            #print(type(top3_scores))
            #print(top3_scores.head())  # Print first few rows
            
            #print(top3_scores.columns)  # List available columns
            #print(top3_scores.index)    # List available row labels


            def highlight_top_values(val, col_name, df, highlight_columns):
                """
                Highlights the top 3 values in a given column.
                """
                if col_name not in highlight_columns:  
                    return ""  # Do nothing if it's not a column we are highlighting
                
                sorted_df = df.sort_values(by=[col_name], ascending=False).reset_index(drop=True)
                top3_values = sorted_df[col_name].head(3).values  # Get top 3 values

                colors1 = {0: '#FDE49B', 1: '#D9D9D9', 2: '#FFE1CC'}  # Bronze for 3rd place
                colors2 = {0: '#FFB3B5', 1: '#FFB3B5', 2: '#FFB3B5'}  # Light red for bottom 3

                if val in top3_values and (col_name != 'Score Against' or col_name != 'GW Transfer Hit' or col_name != 'Total Transfer Hit'):
                    index = list(top3_values).index(val)  # Find index in top3
                    return f'background-color: {colors1.get(index, "")}'  # Apply highlight
                if val in top3_values and (col_name == 'Score Against' or col_name == 'GW Transfer Hit' or col_name == 'Total Transfer Hit'): 
                    index = list(top3_values).index(val)  # Find index in top3
                    return f'background-color: {colors2.get(index, "")}'  # Apply highlight
                
                return ""  # Keep default formatting

            def highlight_bottom_values(val, col_name, df, highlight_columns):
                """
                Highlights the bottom 3 values in a given column.
                """
                if col_name not in highlight_columns:  
                    return ""  # Do nothing if it's not a column we are highlighting
                
                sorted_df = df.sort_values(by=[col_name], ascending=True).reset_index(drop=True)
                bottom3_values = sorted_df[col_name].head(3).values  # Get bottom 3 values

                colors1 = {0: '#FDE49B', 1: '#D9D9D9', 2: '#FFE1CC'}  # Bronze for 3rd place
                colors2 = {0: '#FFB3B5', 1: '#FFB3B5', 2: '#FFB3B5'}  # Light red for bottom 3

                if val in bottom3_values and col_name != 'Score Against':
                    index = list(bottom3_values).index(val)  # Find index in bottom3
                    return f'background-color: {colors2.get(index, "")}'  # Apply highlight
                if val in bottom3_values and col_name == 'Score Against':  
                    index = list(bottom3_values).index(val)  # Find index in top3
                    return f'background-color: {colors1.get(index, "")}'  # Apply highlight
                
                return ""  # Keep default formatting


            
            # highlight top 3 scoring teams
            def highlight_top_scores(val, col_name, t3):
                #if col_name == 'Score' and val in top3_scores['Score'].values:
                #    return 'background-color: yellow'  # or another color to highlight the top scores
                #else:
                #    return ''  # no highlighting for other rows
                if col_name != 'Score':  # Only highlight the "Score" column
                    return ''

                #print(t3.head())  # Debugging: Print first few rows of top 3 scores

                # Define colors for 1st, 2nd, and 3rd place
                colors = {0: '#FDE49B', 1: '#D9D9D9', 2: '#FFE1CC'}  # Bronze hex: #cd7f32

                if val in t3['Score'].values:
                    #i = len(t3) - 1  # Get the last index in t3

                    # Ensure we find the correct index before dropping
                    #print(val)
                    matching_indexes = t3[t3['Score'] == val].index
                    #if len(matching_indexes) > 0:
                    index_to_drop = matching_indexes[0]  # Pick the first matching index
                    #print(index_to_drop)
                        #t3 = t3.drop(index_to_drop).reset_index(drop=True)  # Drop and reset index
                    
                    return f'background-color: {colors.get(index_to_drop, "white")}'  # Default to white if index out of range
                
                return ''  # No highlighting for other rows

            # highlight top 3 scoring teams
            def highlight_bottom_scores(val, col_name, b3):
                #if col_name == 'Score' and val in top3_scores['Score'].values:
                #    return 'background-color: yellow'  # or another color to highlight the top scores
                #else:
                #    return ''  # no highlighting for other rows
                if col_name != 'Score':  # Only highlight the "Score" column
                    return ''

                #print(b3.head())  # Debugging: Print first few rows of top 3 scores

                # Define colors for 1st, 2nd, and 3rd place
                colors = {0: '#FFB3B5', 1: '#FFB3B5', 2: '#FFB3B5'}  # Bronze hex: #cd7f32

                if val in b3['Score'].values:
                    #i = len(t3) - 1  # Get the last index in t3

                    # Ensure we find the correct index before dropping
                    #print(val)
                    matching_indexes = b3[b3['Score'] == val].index
                    #if len(matching_indexes) > 0:
                    index_to_drop = matching_indexes[0]  # Pick the first matching index
                    #print(index_to_drop)
                        #t3 = t3.drop(index_to_drop).reset_index(drop=True)  # Drop and reset index
                    
                    return f'background-color: {colors.get(index_to_drop, "white")}'  # Default to white if index out of range
                
                return ''  # No highlighting for other rows

            # Adjust column width for all columns, accounting for merged cells
            def adjust_column_width(sheet): 
                for col in sheet.columns:
                    max_length = 0
                    column = col[0]  # Start with the first cell in the column
                    
                    # Check if the cell is merged, if it is, get the top-left corner of the merged region
                    if isinstance(column, MergedCell):
                        # Get the top-left corner of the merged cell block
                        merged_cells = sheet.merged_cells
                        for merged_range in merged_cells:
                            if column.coordinate in merged_range:
                                column = sheet[merged_range.min_row][merged_range.min_col - 1]  # Top-left of merged block
                                break
                    
                    column_letter = get_column_letter(column.column)
                    
                    # Now calculate the maximum length of data in the column
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)  # Adding some padding to the width
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # highlight rules for chips
            #styled_primary_df = primary_df.style.map(self.color_chip_background, subset=pd.IndexSlice[:, chip_labels_list])
            
            # highlight rules for team value
            #styled_primary_df = styled_primary_df.map(self.color_team_value_background, subset=pd.IndexSlice[:, 'Current Team Value'])
            
            # highlight rules for change in team value
            #styled_primary_df = styled_primary_df.map(self.color_delta_background, subset=pd.IndexSlice[:, 'Week Value Change'])
            
            # highlight rules for top 3 and bottom 3 scoring teams
            #styled_primary_df = primary_df.style.map(lambda val: highlight_top_scores(val, 'Score'), subset=['Score'])
            #styled_primary_df = primary_df.style.map(lambda val: highlight_top_scores(val, 'Score', top3_scores), subset=['Score'])
            #styled_primary_df = primary_df.style.map(lambda val: highlight_bottom_scores(val, 'Score', bottom3_scores), subset=['Score'])

            # apply all rules at once
            # Apply styles to DataFrame
            #styled_primary_df = primary_df.style \
            #    .map(self.color_chip_background, subset=pd.IndexSlice[:, chip_labels_list]) \
            #    .map(self.color_team_value_background, subset=pd.IndexSlice[:, 'Current Team Value']) \
            #    .map(lambda val: highlight_top_scores(val, 'Score', top3_scores), subset=['Score']) \
            #    .map(lambda val: highlight_bottom_scores(val, 'Score', bottom3_scores), subset=['Score'])
                
            highlight_columns = ['Score', 'Score Against' ,'Plus/Minus', 'GW Points on Bench', 'Season Points on Bench', 'GW Transfers', 'GW Transfer Hit', 'Total Transfers Made', 'Total Transfer Hit' ]  # Columns to highlight

            styled_primary_df = primary_df.style \
                .map(self.color_chip_background, subset=pd.IndexSlice[:, chip_labels_list]) \
                .map(self.color_team_value_background, subset=pd.IndexSlice[:, 'Current Team Value']) \
                .map(lambda val: highlight_top_values(val, 'Score', primary_df, highlight_columns), subset=['Score']) \
                .map(lambda val: highlight_top_values(val, 'Score Against', primary_df, highlight_columns), subset=['Score Against']) \
                .map(lambda val: highlight_top_values(val, 'Plus/Minus', primary_df, highlight_columns), subset=['Plus/Minus']) \
                .map(lambda val: highlight_top_values(val, 'GW Points on Bench', primary_df, highlight_columns), subset=['GW Points on Bench']) \
                .map(lambda val: highlight_top_values(val, 'Season Points on Bench', primary_df, highlight_columns), subset=['Season Points on Bench']) \
                .map(lambda val: highlight_top_values(val, 'GW Transfers', primary_df, highlight_columns), subset=['GW Transfers']) \
                .map(lambda val: highlight_top_values(val, 'Total Transfers Made', primary_df, highlight_columns), subset=['Total Transfers Made']) \
                .map(lambda val: highlight_top_values(val, 'Total Transfer Hit', primary_df, highlight_columns), subset=['Total Transfer Hit']) \
                .map(lambda val: highlight_bottom_values(val, 'Score', primary_df, highlight_columns), subset=['Score']) \
                .map(lambda val: highlight_bottom_values(val, 'Score Against', primary_df, highlight_columns), subset=['Score Against']) \
                .map(lambda val: highlight_bottom_values(val, 'Plus/Minus', primary_df, highlight_columns), subset=['Plus/Minus']) \
                .map(lambda val: highlight_bottom_values(val, 'Season Points on Bench', primary_df, highlight_columns), subset=['Season Points on Bench']) \
                .map(lambda val: highlight_bottom_values(val, 'Total Transfers Made', primary_df, highlight_columns), subset=['Total Transfers Made'])     



            # write out the whole sheet
            styled_primary_df.to_excel(excel_writer=writer, sheet_name=gameweek_str, startcol=3,startrow=1, index=False)

            # Write the secondary table
            sliced_df = primary_df.iloc[:, :11]
            sliced_df.sort_values(by=['Score', 'Plus/Minus'], ascending=[False, False], inplace=True)
            highlight_columns_2 = ['Score', 'Score Against' ,'Plus/Minus']  # Columns to highlight
            styled_secondary_df = sliced_df.style \
                .map(lambda val: highlight_top_values(val, 'Score', sliced_df, highlight_columns_2), subset=['Score']) \
                .map(lambda val: highlight_top_values(val, 'Score Against', sliced_df, highlight_columns_2), subset=['Score Against']) \
                .map(lambda val: highlight_top_values(val, 'Plus/Minus', sliced_df, highlight_columns_2), subset=['Plus/Minus']) \
                .map(lambda val: highlight_bottom_values(val, 'Score', sliced_df, highlight_columns_2), subset=['Score']) \
                .map(lambda val: highlight_bottom_values(val, 'Score Against', sliced_df, highlight_columns_2), subset=['Score Against']) \
                .map(lambda val: highlight_bottom_values(val, 'Plus/Minus', sliced_df, highlight_columns_2), subset=['Plus/Minus'])
            
            styled_secondary_df.to_excel(excel_writer=writer, sheet_name=gameweek_str, startcol=3, startrow=32, index=False)

            # write out fdr table
            #print("fdr")
            #print(fdr_df.shape)  # This will give you (number_of_rows, number_of_columns)
            #print(fdr_df.head())  # This will show you the first few rows of the DataFrame
            #styled_fdr = fdr_df.style.map(self.color_fdr_background)
            #styled_fdr.to_excel(excel_writer=writer, sheet_name=gameweek_str, startcol=15, startrow=32, index=False)

        # TODO: Reach into the previous Gameweek sheet to find the difference in poll position

            # Usage
            wb = openpyxl.load_workbook(results_filepath)
            sheet = wb.active  # Or specify a sheet name like wb['Sheet1']

            adjust_column_width(sheet)

            # Save the workbook with adjusted column widths
            wb.save(results_filepath)
            
            # --- Also emit normalized JSON for the website ---
            # Sort the primary table like the sheet does (Points desc, then Score desc)
            primary_for_json = primary_df.copy()
            primary_for_json.sort_values(by=['Points', 'Score'], ascending=[False, False], inplace=True)

            json_payload = {
                "league": self.file_prefix,          # e.g. "premier" or "championship"
                "sheet": f"GW{gameweek}",
                "generated_at": time.time(),
                "rows": primary_for_json.to_dict(orient="records"),
            }

            json_path = os.path.join(self._latest_dir, f"{self.file_prefix}.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(json_payload, f, ensure_ascii=False)

    

    def create_fdr_data_list(self, data: dict) -> List:
        fdr_list = data['fdr_list']
        name = data['name']
        rating = data['fdr_rating']
        list_data = [f'{name} ({rating})', data['owner']]
        list_data.extend(list(chain.from_iterable(fdr_list)))
        return list_data

    def create_primary_data_list(self, data: dict, current_gw: int = 38) -> List:
        list_data = []
        list_data.append(data['name'])
        list_data.append(data['owner'])

        wins = data['record']['wins']
        draws = data['record']['draws']
        losses = data['record']['losses']

        current_points = 3*wins + draws
        list_data.append(current_points)
        list_data.append(wins)
        list_data.append(draws)
        list_data.append(losses)

        games_played = wins + draws + losses
        games_outstanding = 38 - games_played
        list_data.append(games_played)
        list_data.append(games_outstanding)

        campaign_score = data['campaign_data']['field_score']
        campaign_score_against = data['campaign_data']['field_score_against']
        list_data.append(campaign_score)
        list_data.append(campaign_score_against)
        list_data.append(campaign_score - campaign_score_against)

        # Position difference
        #list_data.append('-')

        list_data.append(data['gameweek_data']['bench_score'])
        list_data.append(data['campaign_data']['bench_score'])

        list_data.append(data['gameweek_data']['transfers'])
        list_data.append(data['gameweek_data']['transfers_cost'])
        list_data.append(data['campaign_data']['transfers'])
        list_data.append(data['campaign_data']['transfers_cost'])

        list_data.append(current_points + (games_outstanding * 3))
        team_value = data['team_value'] / 10
        list_data.append(f'{team_value:.1f}')
        #team_value_delta = 0 #data['gameweek_data']['team_value_delta'] / 10
        #list_data.append(f'{team_value_delta:.1f}')

        tc1 = self._get_chip_str(data['chips']['3xc_1'],     current_gw, half=1)
        bb1 = self._get_chip_str(data['chips']['bboost_1'],  current_gw, half=1)
        fh1 = self._get_chip_str(data['chips']['freehit_1'], current_gw, half=1)
        wc1 = self._get_chip_str(data['chips']['wildcard_1'],current_gw, half=1)
        tc2 = self._get_chip_str(data['chips']['3xc_2'],     current_gw, half=2)
        bb2 = self._get_chip_str(data['chips']['bboost_2'],  current_gw, half=2)
        fh2 = self._get_chip_str(data['chips']['freehit_2'], current_gw, half=2)
        wc2 = self._get_chip_str(data['chips']['wildcard_2'],current_gw, half=2)

        list_data.append(tc1)
        list_data.append(bb1)
        list_data.append(fh1)
        list_data.append(wc1)
        list_data.append(tc2)
        list_data.append(bb2)
        list_data.append(fh2)
        list_data.append(wc2)

        return list_data

    def _get_chip_str(self, week_usage: int, current_gw: int = 38, half: int = 0) -> str:
        """
        half=1 -> first-half chip (valid GW1-19)
        half=2 -> second-half chip (valid GW20-38)
        half=0 -> single chip (manager/assman), always available until used
        """
        HALF_SPLIT = 19
        if week_usage != 0:
            return f'GW{week_usage}'
        # Unused chip
        if half == 1 and current_gw > HALF_SPLIT:
            return 'Expired'
        if half == 2 and current_gw <= HALF_SPLIT:
            return 'Not Yet'
        return 'Available'
